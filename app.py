import math
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def workbook_process(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        after_tax_value = cell.value * 1.1
        after_tax_value_cell = sheet.cell(row, 4)
        after_tax_value_cell.value = after_tax_value
        # print((math.floor(after_tax_value * 100)) / 100.0)

    chart_values = Reference(sheet,
                             min_row=2,
                             max_row=sheet.max_row,
                             min_col=4,
                             max_col=4)

    chart = BarChart()
    chart.add_data(chart_values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


workbook_process('transactions.xlsx')
